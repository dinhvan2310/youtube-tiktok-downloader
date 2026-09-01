from __future__ import annotations

import shutil
import threading
import uuid
import json
import time
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path
from typing import Any
from urllib.parse import urlsplit, urlunsplit

from fastapi import FastAPI, HTTPException
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, Field

from app.config_models import AppConfig, SourceConfig, validate_global_config, validate_source
from app.config_store import default_config_path, load_config, save_config
from app.db import DownloadDB
from app.filename import count_videos_in_folder
from app.orchestrator import DownloadOrchestrator

app = FastAPI(title="Tool Download Video API", version="1.0.0")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])
_jobs: dict[str, dict[str, Any]] = {}
_lock = threading.Lock()
_executor = ThreadPoolExecutor(max_workers=16)
_lane_condition = threading.Condition()
_lane_running = {"crawl": 0, "download": 0, "metadata": 0}
DownloadDB().interrupt_unfinished_jobs()


class ConfigPayload(BaseModel):
    quality_height: int = 0
    target_videos_per_page: int = 20
    source_thread_count: int = 4
    download_source_concurrency: int = 4
    global_download_concurrency: int = 12
    metadata_workers: int = 6
    source_root_path: str = ""
    youtube_cookies_file: str = ""
    youtube_cookies_browser: str = ""
    tiktok_cookies_file: str = ""
    sources: list[dict[str, Any]] = Field(default_factory=list)


class JobPayload(BaseModel):
    kind: str
    source_ids: list[str] | None = None
    batch_size: int = 50
    resume: bool = False


class FilePayload(BaseModel):
    path: str
    root_path: str = ""


class SourcePayload(BaseModel):
    source_id: str = ""
    path_download: str
    links: list[str] = Field(default_factory=list)
    note: str = ""
    reup_source: str = "youtube_tiktok"
    move_duplicate_links: bool = False


class VideoRef(BaseModel):
    source_id: str
    video_id: str


class VideoTransitionPayload(BaseModel):
    items: list[VideoRef]
    status: str
    expected_status: str | None = None
    reason: str = ""
    note: str = ""
    review_at: str = ""


def _lane_for(kind: str) -> str:
    return "crawl" if kind in {"crawl", "crawl-batched"} else "metadata" if kind == "metadata" else "download"


def _lane_limit(kind: str, cfg: AppConfig) -> int:
    lane = _lane_for(kind)
    if lane == "crawl": return cfg.global_config.source_thread_count
    if lane == "metadata": return cfg.global_config.metadata_workers
    return cfg.global_config.download_source_concurrency


def _acquire_lane(kind: str, cfg: AppConfig) -> str:
    lane = _lane_for(kind)
    with _lane_condition:
        while _lane_running[lane] >= _lane_limit(kind, cfg):
            _lane_condition.wait(timeout=1)
        _lane_running[lane] += 1
    return lane


def _release_lane(lane: str) -> None:
    with _lane_condition:
        _lane_running[lane] = max(0, _lane_running[lane] - 1)
        _lane_condition.notify_all()


def _config() -> AppConfig:
    return load_config()


def _canonical_link(value: str) -> str:
    parts = urlsplit(value.strip())
    host = parts.netloc.lower().removeprefix("www.")
    path = parts.path.rstrip("/")
    if host == "youtube.com" and path.endswith("/videos"):
        path = path[:-7]
    return urlunsplit(("https", host, path, "", ""))


def _active_sources(cfg: AppConfig, db: DownloadDB | None = None) -> list[SourceConfig]:
    statuses = (db or DownloadDB()).get_source_statuses()
    return [source for source in cfg.sources if statuses.get(source.id, "active") == "active"]


def _job_log(job_id: str, message: str) -> None:
    with _lock:
        job = _jobs.get(job_id)
        if job:
            job["logs"].append(message)
            job["logs"] = job["logs"][-2000:]
    DownloadDB().add_job_event(job_id, "log", message=message, level="error" if message.startswith("ERROR") else "info")


def _job_event(job_id: str, event: str, data: dict[str, Any]) -> None:
    with _lock:
        job = _jobs.get(job_id)
        if job:
            job.setdefault("events", []).append({"event": event, "data": data})
            job["events"] = job["events"][-1000:]
    failed = event == "failed" or event.endswith("_failed")
    message = str(data.get("error") or "")
    if event.startswith("download_") and not message:
        title = str(data.get("title") or data.get("source_label") or "Download")
        percent = data.get("percent")
        message = f"{title}{f' — {percent}%' if percent is not None else ''}"
    DownloadDB().add_job_event(job_id, event, message=message, level="error" if failed else "info", data=data)


def _run_job(job_id: str, payload: JobPayload) -> None:
    cfg, db = _config(), DownloadDB()
    lane = _acquire_lane(payload.kind, cfg)
    with _lock:
        stop = _jobs[job_id].get("stop_event") or threading.Event()
        _jobs[job_id].update(stop_event=stop, status="stopping" if stop.is_set() else "running")
    db.update_job_record(job_id, "running")
    _job_event(job_id, "started", {"kind": payload.kind, "source_ids": payload.source_ids or []})
    sources = _active_sources(cfg, db)
    if payload.source_ids is not None:
        selected = set(payload.source_ids)
        sources = [s for s in sources if s.id in selected]
    try:
        if payload.kind in {"crawl", "crawl-batched", "download"}:
            runner = DownloadOrchestrator(cfg, log=lambda m: _job_log(job_id, m),
                progress=lambda m: _job_log(job_id, m), progress_log=lambda m: _job_log(job_id, m),
                progress_event=lambda event, data: _job_event(job_id, event, data))
            runner._stop_event = stop
            if payload.kind == "crawl-batched":
                def on_batch(source, sequence, items):
                    _job_event(job_id, "batch_ready", {"source_id": source.id, "sequence": sequence, "items": items, "count": len(items)})
                result = runner.crawl_sources_batched(sources, batch_size=max(1, payload.batch_size), on_batch=on_batch, start_batch=-1 if payload.resume else 0)
            elif payload.kind == "crawl": result = runner.crawl_sources(sources)
            else: result = runner.download_sources(sources)
            summary = {"crawled": result.crawled, "downloaded": result.downloaded, "skipped": result.skipped,
                       "errors": result.errors, "sources_done": result.sources_done}
        elif payload.kind == "metadata":
            from app.playlist import fetch_video_metadata
            from concurrent.futures import as_completed
            candidates = db.metadata_candidates([s.id for s in sources], limit=500)
            enriched = failed = 0
            def enrich(row):
                try:
                    data = fetch_video_metadata(row["source_url"], cfg.global_config)
                    db.update_video_metadata(row["source_id"], row["video_id"], data)
                    return True
                except Exception:
                    db.update_video_metadata(row["source_id"], row["video_id"], {}, status="missing")
                    return False
            with ThreadPoolExecutor(max_workers=cfg.global_config.metadata_workers) as pool:
                futures = [pool.submit(enrich, row) for row in candidates]
                for index, future in enumerate(as_completed(futures), 1):
                    if stop.is_set(): break
                    if future.result(): enriched += 1
                    else: failed += 1
                    if index % 10 == 0 or index == len(futures):
                        _job_event(job_id, "progress", {"completed": index, "total": len(futures), "enriched": enriched})
            summary = {"enriched": enriched, "errors": failed, "sources_done": len(sources)}
        else:
            raise ValueError(f"Unknown job kind: {payload.kind}")
        final_status = "stopped" if stop.is_set() else "done"
        with _lock:
            _jobs[job_id].update(status=final_status, result=summary)
        db.update_job_record(job_id, final_status, result=summary)
        _job_event(job_id, "stopped" if final_status == "stopped" else "completed", summary)
        if final_status == "done" and payload.kind in {"crawl", "crawl-batched"} and sources:
            try:
                create_job(JobPayload(kind="metadata", source_ids=[source.id for source in sources]))
            except HTTPException:
                pass
    except Exception as exc:
        _job_log(job_id, f"ERROR: {exc}")
        with _lock:
            _jobs[job_id].update(status="failed", error=str(exc))
        db.update_job_record(job_id, "failed", error=str(exc))
        _job_event(job_id, "failed", {"error": str(exc)})
    finally:
        _release_lane(lane)


@app.get("/api/health")
def health() -> dict[str, str]:
    return {"status": "ok"}


@app.get("/api/config")
def get_config() -> dict[str, Any]:
    return _config().to_dict()


@app.put("/api/config")
def put_config(payload: ConfigPayload) -> dict[str, Any]:
    data = payload.model_dump() if hasattr(payload, "model_dump") else payload.dict()
    cfg = AppConfig.from_dict(data)
    errors = validate_global_config(cfg.global_config)
    statuses = DownloadDB().get_source_statuses()
    for source in cfg.sources:
        if statuses.get(source.id, "active") != "archived":
            errors.extend(validate_source(source))
    if errors:
        raise HTTPException(400, detail=errors)
    save_config(cfg)
    return cfg.to_dict()


@app.get("/api/sources/stats")
def source_stats(include_archived: bool = False) -> list[dict[str, Any]]:
    cfg, db = _config(), DownloadDB()
    aggregates = db.get_source_overview_counts()
    statuses = db.get_source_statuses()
    result = []
    for source in cfg.sources:
        lifecycle = statuses.get(source.id, "active")
        if lifecycle == "archived" and not include_archived:
            continue
        history = db.get_source_stats(source.id)
        aggregate = aggregates.get(str(source.id), {})
        stamp = history.get("last_crawl_at")
        try:
            age = time.time() - datetime.fromisoformat(str(stamp).replace("Z", "+00:00")).timestamp() if stamp else float("inf")
        except (TypeError, ValueError, OverflowError):
            age = float("inf")
        result.append({**source.to_dict(), "label": source.note or Path(source.path_download).name,
            "crawled_count": aggregate.get("discovered_count", 0),
            "discovered_count": aggregate.get("discovered_count", 0),
            "review_count": aggregate.get("review_count", 0),
            "queued_count": aggregate.get("queued_count", 0),
            "downloaded_count": aggregate.get("downloaded_count", 0),
            "folder_count": count_videos_in_folder(Path(source.path_download)),
            "capacity": max(0, cfg.global_config.target_videos_per_page - count_videos_in_folder(Path(source.path_download))),
            "status": lifecycle, "last_crawl_at": stamp, "stale": age > 15 * 60})
    return result


def _duplicate_links(cfg: AppConfig, links: list[str], *, exclude_source_id: str = "") -> list[dict[str, str]]:
    statuses = DownloadDB().get_source_statuses()
    owners: dict[str, SourceConfig] = {}
    for source in cfg.sources:
        if source.id == exclude_source_id or statuses.get(source.id, "active") == "archived":
            continue
        for link in source.links:
            owners[_canonical_link(link)] = source
    result = []
    for link in links:
        owner = owners.get(_canonical_link(link))
        if owner:
            result.append({"link": link, "source_id": owner.id, "source_label": owner.note or Path(owner.path_download).name})
    return result


@app.post("/api/sources/validate")
def validate_source_payload(payload: SourcePayload) -> dict[str, Any]:
    candidate = SourceConfig(id="preview", path_download=payload.path_download, links=payload.links,
                             note=payload.note, reup_source=payload.reup_source)
    errors = validate_source(candidate)
    duplicates = _duplicate_links(_config(), candidate.links, exclude_source_id=payload.source_id)
    return {"valid": not errors and not duplicates, "errors": errors, "duplicates": duplicates,
            "preview": {"label": candidate.note or Path(candidate.path_download).name or "Source",
                        "links": [{"url": link, "canonical": _canonical_link(link)} for link in candidate.links]}}


@app.post("/api/sources")
def create_source(payload: SourcePayload) -> dict[str, Any]:
    cfg, db = _config(), DownloadDB()
    source = SourceConfig(id=cfg.new_source_id(), path_download=payload.path_download, links=payload.links,
                          note=payload.note, reup_source=payload.reup_source)
    errors = validate_source(source)
    duplicates = _duplicate_links(cfg, source.links)
    if errors:
        raise HTTPException(400, detail=errors)
    if duplicates and not payload.move_duplicate_links:
        raise HTTPException(409, detail={"message": "Channel link already belongs to another active source", "duplicates": duplicates})
    if duplicates:
        duplicate_links = {_canonical_link(item["link"]) for item in duplicates}
        emptied_source_ids = []
        for current in cfg.sources:
            current.links = [link for link in current.links if _canonical_link(link) not in duplicate_links]
            if not current.links:
                emptied_source_ids.append(current.id)
    cfg.sources.append(source)
    save_config(cfg)
    for emptied_source_id in locals().get("emptied_source_ids", []):
        db.set_source_status(emptied_source_id, "archived")
    db.set_source_status(source.id, "active")
    return source.to_dict()


@app.patch("/api/sources/{source_id}")
def update_source(source_id: str, payload: SourcePayload) -> dict[str, Any]:
    cfg, db = _config(), DownloadDB(); current = next((source for source in cfg.sources if source.id == source_id), None)
    if not current:
        raise HTTPException(404, detail="Source not found")
    candidate = SourceConfig(id=source_id, path_download=payload.path_download, links=payload.links,
                             note=payload.note, reup_source=payload.reup_source)
    errors = validate_source(candidate); duplicates = _duplicate_links(cfg, candidate.links, exclude_source_id=source_id)
    if errors:
        raise HTTPException(400, detail=errors)
    if duplicates and not payload.move_duplicate_links:
        raise HTTPException(409, detail={"message": "Channel link already belongs to another active source", "duplicates": duplicates})
    if duplicates:
        duplicate_links = {_canonical_link(item["link"]) for item in duplicates}
        emptied_source_ids = []
        for source in cfg.sources:
            if source.id != source_id:
                source.links = [link for link in source.links if _canonical_link(link) not in duplicate_links]
                if not source.links:
                    emptied_source_ids.append(source.id)
    cfg.sources = [candidate if source.id == source_id else source for source in cfg.sources]
    save_config(cfg)
    for emptied_source_id in locals().get("emptied_source_ids", []):
        db.set_source_status(emptied_source_id, "archived")
    db.set_source_status(source_id, "active")
    return candidate.to_dict()


@app.post("/api/sources/{source_id}/archive")
def archive_source(source_id: str) -> dict[str, str]:
    cfg = _config()
    if not any(source.id == source_id for source in cfg.sources):
        raise HTTPException(404, detail="Source not found")
    DownloadDB().set_source_status(source_id, "archived")
    with _lock:
        active = next((job for job in _jobs.values() if job.get("status") == "running" and source_id in job.get("source_ids", [])), None)
        if active and active.get("stop_event"):
            active["stop_event"].set(); active["status"] = "stopping"
    return {"status": "archived"}


@app.post("/api/sources/{source_id}/restore")
def restore_source(source_id: str) -> dict[str, str]:
    cfg = _config()
    if not any(source.id == source_id for source in cfg.sources):
        raise HTTPException(404, detail="Source not found")
    DownloadDB().set_source_status(source_id, "active")
    return {"status": "active"}


@app.delete("/api/sources/{source_id}")
def delete_source(source_id: str) -> dict[str, Any]:
    cfg = _config()
    source = next((item for item in cfg.sources if item.id == source_id), None)
    if not source:
        raise HTTPException(404, detail="Source not found")
    status = DownloadDB().get_source_statuses().get(source_id, "active")
    if status != "archived":
        raise HTTPException(409, detail="Archive this source before deleting it permanently")
    with _lock:
        active = next((job for job in _jobs.values() if job.get("status") in {"queued", "running", "stopping"} and source_id in job.get("source_ids", [])), None)
        if active:
            raise HTTPException(409, detail="Wait for the source task to stop before deleting it")
    cfg.sources = [item for item in cfg.sources if item.id != source_id]
    save_config(cfg)
    deleted = DownloadDB().delete_source_data(source_id)
    return {"status": "deleted", "source_id": source_id, "files_removed": False, "records_removed": sum(deleted.values()), "details": deleted}


@app.post("/api/jobs")
def create_job(payload: JobPayload) -> dict[str, str]:
    cfg = _config()
    source_ids = payload.source_ids or [source.id for source in _active_sources(cfg)]
    lane = _lane_for(payload.kind)
    source_set = set(source_ids)
    active = [job for job in _jobs.values() if job["status"] in {"queued", "running", "stopping"}]
    for job in active:
        if _lane_for(job["kind"]) == lane and source_set.intersection(job.get("source_ids", [])):
            raise HTTPException(409, detail=f"A {lane} task is already active for this source")
    if lane == "download":
        paths = {str(Path(source.path_download).resolve()).lower() for source in cfg.sources if source.id in source_set}
        for job in active:
            if _lane_for(job["kind"]) != "download": continue
            active_paths = {str(Path(source.path_download).resolve()).lower() for source in cfg.sources if source.id in job.get("source_ids", [])}
            if paths.intersection(active_paths):
                raise HTTPException(409, detail="A download task is already using this destination folder")
    job_id = uuid.uuid4().hex
    with _lock:
        _jobs[job_id] = {"id": job_id, "kind": payload.kind, "status": "queued", "source_ids": source_ids,
                         "started_at": time.time(), "logs": [], "events": [], "result": None, "stop_event": threading.Event()}
    DownloadDB().create_job_record(job_id, payload.kind, source_ids, status="queued")
    _executor.submit(_run_job, job_id, payload)
    return {"id": job_id}


@app.get("/api/jobs")
def list_jobs(limit: int = 50) -> list[dict[str, Any]]:
    return DownloadDB().list_jobs(limit)


@app.get("/api/overview")
def overview() -> dict[str, Any]:
    cfg, db = _config(), DownloadDB()
    counts = db.get_overview_counts()
    now = time.time()
    freshness = 15 * 60
    source_sync = []
    statuses = db.get_source_statuses()
    for source in cfg.sources:
        if statuses.get(source.id, "active") == "archived":
            continue
        history = db.get_source_stats(source.id)
        stamp = history.get("last_crawl_at")
        try:
            age = now - datetime.fromisoformat(str(stamp).replace("Z", "+00:00")).timestamp() if stamp else float("inf")
        except (TypeError, ValueError, OverflowError):
            age = float("inf")
        source_sync.append({"id": source.id, "label": source.note or Path(source.path_download).name or "Source", "last_crawl_at": stamp, "stale": age > freshness})
    with _lock:
        active_jobs = [{
            "id": j["id"], "kind": j.get("kind"), "status": j["status"],
            "started_at": j.get("started_at"), "log_tail": j.get("logs", [])[-1:]}
            for j in _jobs.values() if j.get("status") in {"queued", "running", "stopping"}]
    return {"sources": len(source_sync), "counts": counts, "active_job": active_jobs[0] if active_jobs else None,
            "active_jobs": active_jobs,
            "last_sync": max((s["last_crawl_at"] for s in source_sync if s["last_crawl_at"]), default=None),
            "stale_sources": [s for s in source_sync if s["stale"]], "sync_errors": counts["failed"]}


@app.post("/api/sync/ensure")
def ensure_sync() -> dict[str, Any]:
    with _lock:
        active_sources = {sid for job in _jobs.values() if _lane_for(job.get("kind", "")) == "crawl" and job.get("status") in {"queued", "running", "stopping"} for sid in job.get("source_ids", [])}
    cfg, db = _config(), DownloadDB()
    now = time.time(); stale: list[str] = []
    statuses = db.get_source_statuses()
    for source in cfg.sources:
        if statuses.get(source.id, "active") != "active":
            continue
        stamp = db.get_source_stats(source.id).get("last_crawl_at")
        try:
            age = now - datetime.fromisoformat(str(stamp).replace("Z", "+00:00")).timestamp() if stamp else float("inf")
        except (TypeError, ValueError, OverflowError):
            age = float("inf")
        if age > 15 * 60 and source.id not in active_sources:
            stale.append(source.id)
    if not stale:
        active_source_ids = [source.id for source in _active_sources(cfg, db)]
        with _lock:
            metadata_active = any(_lane_for(job.get("kind", "")) == "metadata" and job.get("status") in {"queued", "running", "stopping"} for job in _jobs.values())
        if not metadata_active and db.metadata_candidates(active_source_ids, limit=1):
            result = create_job(JobPayload(kind="metadata", source_ids=active_source_ids))
            return {"status": "metadata_started", "job_id": result["id"], "job_ids": [result["id"]], "source_ids": active_source_ids}
        return {"status": "fresh", "job_id": None, "job_ids": [], "source_ids": []}
    job_ids = [create_job(JobPayload(kind="crawl-batched", source_ids=[source_id], batch_size=50, resume=False))["id"] for source_id in stale]
    return {"status": "started", "job_id": job_ids[0] if job_ids else None, "job_ids": job_ids, "source_ids": stale}


@app.get("/api/jobs/{job_id}")
def get_job(job_id: str) -> dict[str, Any]:
    with _lock:
        job = _jobs.get(job_id)
        if not job: raise HTTPException(404, detail="Job not found")
        return {k: v for k, v in job.items() if k != "stop_event"}


@app.get("/api/jobs/{job_id}/events")
def job_events(job_id: str):
    with _lock:
        if job_id not in _jobs:
            raise HTTPException(404, detail="Job not found")
    def stream():
        cursor = 0
        while True:
            with _lock:
                job = _jobs.get(job_id)
                if not job:
                    return
                events = list(job.get("events", []))
                status = job.get("status")
            while cursor < len(events):
                item = events[cursor]; cursor += 1
                yield f"event: {item['event']}\ndata: {json.dumps(item['data'], ensure_ascii=False)}\n\n"
            if status in {"done", "failed", "stopped"} and cursor >= len(events):
                return
            yield ": keep-alive\n\n"
            time.sleep(0.5)
    return StreamingResponse(stream(), media_type="text/event-stream", headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


@app.get("/api/crawled/{source_id}")
def list_crawled(source_id: str, keyword: str = "", status: str | None = None) -> list[dict[str, Any]]:
    rows = DownloadDB().list_crawled(source_id, keyword=keyword, status=status)
    return [dict(row) for row in rows]


@app.get("/api/videos")
def list_videos(status: str = "discovered", source_id: str = "", keyword: str = "",
                exclude_keyword: str = "", min_duration: float | None = None, max_duration: float | None = None,
                min_views: int | None = None, published_after: str = "", published_before: str = "",
                metadata_status: str = "", content_type: str = "", sort: str = "newest",
                limit: int = 200, offset: int = 0) -> dict[str, Any]:
    cfg = _config(); labels = {source.id: source.note or Path(source.path_download).name for source in cfg.sources}
    rows = DownloadDB().list_videos_global(status=status, source_id=source_id, keyword=keyword,
        exclude_keyword=exclude_keyword, min_duration=min_duration, max_duration=max_duration,
        min_views=min_views, published_after=published_after, published_before=published_before,
        metadata_status=metadata_status, content_type=content_type, sort=sort, limit=limit, offset=offset)
    items = []
    for row in rows:
        item = {**dict(row), "source_label": labels.get(str(row["source_id"]), "Source")}
        if item.get("platform") == "youtube" and not item.get("thumbnail"):
            item["thumbnail"] = f"https://i.ytimg.com/vi/{item['video_id']}/hqdefault.jpg"
        items.append(item)
    return {"items": items, "limit": limit, "offset": offset, "has_more": len(items) == min(max(limit, 1), 500)}


@app.post("/api/videos/transition")
def transition_videos(payload: VideoTransitionPayload) -> dict[str, int]:
    allowed = {"discovered", "queued", "ignored", "hold"}
    if payload.status not in allowed:
        raise HTTPException(400, detail="Invalid video status")
    changed = DownloadDB().transition_videos(
        [(item.source_id, item.video_id) for item in payload.items],
        status=payload.status, expected_status=payload.expected_status, reason=payload.reason,
        note=payload.note, review_at=payload.review_at,
    )
    return {"updated": changed}


@app.get("/api/queue")
def list_queue(limit: int = 500) -> dict[str, Any]:
    cfg = _config(); db = DownloadDB()
    sources = {source.id: source for source in cfg.sources}
    rows = db.list_videos_global(status="queued", limit=limit, exclude_downloaded=True)
    items = []
    per_source_ready: dict[str, int] = {}
    for row in rows:
        source_id = str(row["source_id"]); source = sources.get(source_id)
        if not source:
            continue
        capacity = max(0, cfg.global_config.target_videos_per_page - count_videos_in_folder(Path(source.path_download)))
        used = per_source_ready.get(source_id, 0)
        queue_state = "ready" if used < capacity else "waiting_capacity"
        if queue_state == "ready":
            per_source_ready[source_id] = used + 1
        items.append({**dict(row), "source_label": source.note or Path(source.path_download).name,
                      "path_download": source.path_download, "capacity": capacity, "queue_state": queue_state})
    return {"items": items, "ready": sum(item["queue_state"] == "ready" for item in items),
            "waiting_capacity": sum(item["queue_state"] == "waiting_capacity" for item in items)}


class QueuePayload(BaseModel):
    source_id: str
    video_ids: list[str]
    status: str = "queued"


@app.post("/api/queue/approve")
def approve_queue(payload: QueuePayload) -> dict[str, int]:
    if payload.status not in {"queued", "approved", "shortlisted"}:
        raise HTTPException(400, detail="Invalid queue status")
    return {"updated": DownloadDB().set_crawled_status(payload.source_id, payload.video_ids, payload.status)}


@app.post("/api/jobs/{job_id}/stop")
def stop_job(job_id: str) -> dict[str, str]:
    with _lock:
        job = _jobs.get(job_id)
        if not job: raise HTTPException(404, detail="Job not found")
        event = job.get("stop_event")
        if event: event.set()
        job["status"] = "stopping"
    DownloadDB().update_job_record(job_id, "stopping")
    return {"status": "stopping"}


@app.post("/api/folders/clean")
def clean_folders() -> dict[str, Any]:
    cfg = _config(); root = Path(cfg.global_config.source_root_path)
    if not root.is_dir(): raise HTTPException(400, detail="Root path not found")
    known = {Path(s.path_download).resolve() for s in cfg.sources if s.path_download}
    removed = []
    for item in root.iterdir():
        if item.is_dir() and item.resolve() not in known:
            shutil.rmtree(item)
            removed.append(item.name)
    return {"removed": removed}

def _read_excel_sources(path: str, root_path: str = "") -> tuple[AppConfig, list[SourceConfig], list[str]]:
    import openpyxl
    cfg = _config()
    workbook_path = Path(path)
    if not workbook_path.is_file():
        raise HTTPException(400, detail="Excel file not found")
    root_value = (root_path or cfg.global_config.source_root_path).strip()
    if not root_value:
        raise HTTPException(400, detail="Choose a download root folder for the imported sources")
    root = Path(root_value).expanduser().resolve()
    if not root.is_dir():
        raise HTTPException(400, detail="Download root folder not found. Choose another folder and retry")
    cfg.global_config.source_root_path = str(root)
    try:
        wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception as error:
        raise HTTPException(400, detail=f"Could not read Excel file: {error}") from error
    if not rows: return cfg, [], []
    headers = [str(x or '').strip().lower() for x in rows[0]]
    def col(name: str) -> int | None:
        return headers.index(name) if name in headers else None
    folder_i = col('folder name')
    if folder_i is None: raise HTTPException(400, detail="Excel needs a 'folder name' column")
    link_cols = [i for i, h in enumerate(headers) if h.startswith('link')]
    note_i, reup_i = col('note'), col('nguồn reup')
    incoming: list[SourceConfig] = []; errors: list[str] = []
    for row_number, row in enumerate(rows[1:], start=2):
        if not any(value is not None and str(value).strip() for value in row):
            continue
        folder = str(row[folder_i] or '').strip()
        links = [str(row[i]).strip() for i in link_cols if i < len(row) and row[i]]
        folder_path = Path(folder)
        if not folder or folder_path.is_absolute() or folder in {'.', '..'} or '..' in folder_path.parts:
            errors.append(f"Row {row_number}: invalid folder name '{folder or '(blank)'}'")
            continue
        destination = (root / folder_path).resolve()
        if not destination.is_relative_to(root):
            errors.append(f"Row {row_number}: folder must stay inside the selected root")
            continue
        source = SourceConfig(id='', path_download=str(destination), links=links,
            note=str(row[note_i] or '').strip() if note_i is not None and note_i < len(row) else '',
            reup_source=str(row[reup_i] or '').strip() if reup_i is not None and reup_i < len(row) else '')
        row_errors = validate_source(source)
        if row_errors: errors.extend(f"Row {row_number} ({folder}): {error}" for error in row_errors)
        else: incoming.append(source)
    return cfg, incoming, errors


@app.post("/api/excel/preview")
def preview_excel(payload: FilePayload) -> dict[str, Any]:
    cfg, incoming, errors = _read_excel_sources(payload.path, payload.root_path)
    duplicates = []
    for source in incoming:
        duplicates.extend(_duplicate_links(cfg, source.links))
    return {"rows": len(incoming), "errors": errors, "duplicates": duplicates,
            "root_path": cfg.global_config.source_root_path,
            "preview": [{"folder": source.path_download, "note": source.note, "links": len(source.links)} for source in incoming[:50]]}


@app.post("/api/excel/import")
def import_excel(payload: FilePayload) -> dict[str, Any]:
    cfg, incoming, errors = _read_excel_sources(payload.path, payload.root_path)
    if errors:
        raise HTTPException(400, detail=errors)
    old = {str(Path(s.path_download).resolve()).lower(): s for s in cfg.sources}
    used = {s.id for s in cfg.sources}; result = list(cfg.sources)
    for source in incoming:
        key = str(Path(source.path_download).resolve()).lower(); previous = old.get(key)
        source.id = previous.id if previous else cfg.new_source_id()
        while source.id in used and source.id != (previous.id if previous else None): source.id = cfg.new_source_id()
        used.add(source.id)
        if previous: result = [source if item.id == previous.id else item for item in result]
        else: result.append(source)
    cfg.sources = result; save_config(cfg)
    return cfg.to_dict()

@app.post("/api/excel/export")
def export_excel(payload: FilePayload) -> dict[str, str]:
    import openpyxl
    cfg = _config(); root = cfg.global_config.source_root_path
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'sources'
    ws.append(['folder name', 'link1', 'link2', 'link3', 'link4', 'link5', 'note', 'nguồn reup'])
    for source in cfg.sources:
        path = Path(source.path_download); folder = str(path.relative_to(Path(root))) if root and path.is_relative_to(Path(root)) else path.name
        ws.append([folder, *(source.links[:5] + [''] * 5)[:5], source.note, source.reup_source])
    wb.save(payload.path); return {'path': payload.path}
